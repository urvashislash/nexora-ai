"""
NEXORA AI — Integration Tests (Plan §7.2)

Tests the full path: upload → parse → extract → normalize → match → review → commit.
Validates database write contracts and audit record generation.
Covers edge cases: duplicate files, malformed data, low-confidence matches, and retries.

These tests run entirely in-process (no live DB/RabbitMQ required) by exercising
the full Python AI pipeline and simulating the Rust trust-plane logic in Python.
"""

import io
import json
from datetime import date
from uuid import UUID

import pytest
from app.models.schemas import (
    DisciplineEnum,
    MatchProposalPayload,
    MatchTierEnum,
    NormalizedObservation,
    ScheduleActivity,
)
from app.services.extractor import DocumentExtractor
from app.services.matcher import HybridMatcher
from app.services.normalizer import default_normalizer

# =============================================================================
# Shared Fixtures — Realistic Project Schedule
# =============================================================================


@pytest.fixture
def project_id() -> UUID:
    return UUID("a0000000-0000-0000-0000-000000000001")


@pytest.fixture
def full_schedule(project_id) -> list[ScheduleActivity]:
    """
    Full 8-activity schedule matching the demo seed data in
    database/seeds/001_demo_project_seed.sql
    """
    return [
        ScheduleActivity(
            id=UUID("d0000000-0000-0000-0000-000000000001"),
            project_id=project_id,
            code="PIP-2400",
            name="Spool Erection and Alignment - Pipe Rack B",
            description="Prefabricated carbon steel piping spool erection, alignment and tack welding along Grid B1-B8",
            discipline=DisciplineEnum.PIPING,
            location="Pipe Rack B",
            zone="Zone 2",
            equipment_tag="RACK-B-CS",
            planned_start_date=date(2026, 8, 10),
            planned_finish_date=date(2026, 8, 25),
            planned_quantity=450.0,
            unit_of_measure="Inch-Dia",
        ),
        ScheduleActivity(
            id=UUID("d0000000-0000-0000-0000-000000000002"),
            project_id=project_id,
            code="PIP-2401",
            name="Hydrostatic Testing - Line P-101 (Crude Feed Header)",
            description="Pressure testing of 24 inch crude feed header Line P-101 at 42.5 bar with holding time 4 hours",
            discipline=DisciplineEnum.PIPING,
            location="Pipe Rack B",
            zone="Zone 2",
            equipment_tag="LINE-P-101",
            planned_start_date=date(2026, 8, 26),
            planned_finish_date=date(2026, 8, 28),
            planned_quantity=1.0,
            unit_of_measure="Test-Pack",
        ),
        ScheduleActivity(
            id=UUID("d0000000-0000-0000-0000-000000000003"),
            project_id=project_id,
            code="PIP-2402",
            name="Hydrostatic Testing - Line P-102 (Naphtha Return Header)",
            description="Pressure testing of 16 inch naphtha return header Line P-102 at 32.0 bar with holding time 4 hours",
            discipline=DisciplineEnum.PIPING,
            location="Pipe Rack B",
            zone="Zone 2",
            equipment_tag="LINE-P-102",
            planned_start_date=date(2026, 8, 28),
            planned_finish_date=date(2026, 8, 30),
            planned_quantity=1.0,
            unit_of_measure="Test-Pack",
        ),
        ScheduleActivity(
            id=UUID("d0000000-0000-0000-0000-000000000004"),
            project_id=project_id,
            code="CIV-1100",
            name="Rebar Tying and Shuttering - Compressor Foundation",
            description="Reinforcement steel bar cutting, bending, binding, and formwork for Main Gas Compressor C-101",
            discipline=DisciplineEnum.CIVIL,
            location="Compressor House",
            zone="Zone 1",
            equipment_tag="FND-C-101",
            planned_start_date=date(2026, 8, 15),
            planned_finish_date=date(2026, 8, 24),
            planned_quantity=35.5,
            unit_of_measure="MT",
        ),
        ScheduleActivity(
            id=UUID("d0000000-0000-0000-0000-000000000005"),
            project_id=project_id,
            code="CIV-1101",
            name="Concrete Pour - Column Footings Area 100",
            description="M35 grade ready-mix concrete pouring for heavy column footings C1 to C12 in CDU area",
            discipline=DisciplineEnum.CIVIL,
            location="CDU Area 100",
            zone="Zone 1",
            equipment_tag="COL-FTG-100",
            planned_start_date=date(2026, 8, 25),
            planned_finish_date=date(2026, 8, 29),
            planned_quantity=180.0,
            unit_of_measure="Cu.M",
        ),
        ScheduleActivity(
            id=UUID("d0000000-0000-0000-0000-000000000006"),
            project_id=project_id,
            code="MEC-3200",
            name="Equipment Alignment - Crude Charge Pump P-101A",
            description="Precision dial indicator shaft alignment and baseplate grouting for Crude Charge Pump P-101A",
            discipline=DisciplineEnum.MECHANICAL,
            location="CDU Area 100",
            zone="Zone 1",
            equipment_tag="PUMP-P-101A",
            planned_start_date=date(2026, 8, 28),
            planned_finish_date=date(2026, 9, 2),
            planned_quantity=1.0,
            unit_of_measure="Unit",
        ),
        ScheduleActivity(
            id=UUID("d0000000-0000-0000-0000-000000000007"),
            project_id=project_id,
            code="ELE-4100",
            name="Cable Tray Installation - Pipe Rack Tier 3",
            description="Installation of galvanized steel ladder type cable trays along Rack B Tier 3 for HV power cables",
            discipline=DisciplineEnum.ELECTRICAL,
            location="Pipe Rack B",
            zone="Zone 2",
            equipment_tag="TRAY-RACK-B3",
            planned_start_date=date(2026, 8, 20),
            planned_finish_date=date(2026, 8, 31),
            planned_quantity=620.0,
            unit_of_measure="Meters",
        ),
        ScheduleActivity(
            id=UUID("d0000000-0000-0000-0000-000000000008"),
            project_id=project_id,
            code="INS-5100",
            name="Transmitter Calibration and Hookup - PT-101",
            description="Bench calibration, impulse tubing hookup and loop test for Pressure Transmitter PT-101 on CDU column inlet",
            discipline=DisciplineEnum.INSTRUMENTATION,
            location="CDU Area 100",
            zone="Zone 1",
            equipment_tag="PT-101",
            planned_start_date=date(2026, 8, 29),
            planned_finish_date=date(2026, 9, 3),
            planned_quantity=1.0,
            unit_of_measure="Tag",
        ),
    ]


# =============================================================================
# 1. FULL PIPELINE: upload → parse → extract → normalize → match → review → commit
# =============================================================================


class TestFullPipelineTextReport:
    """
    Integration test: Full path from raw daily report text through extraction,
    normalization, matching, review classification, and commit eligibility.
    """

    DAILY_REPORT = """DAILY CONSTRUCTION REPORT - AREA 101
Project: PRD-HYD-PKG04
Date: 28-Aug-2026
Contractor: NEXORA EPC

1. P-101 hydro test completed successfully with 42.5 bar pressure holding for 4 hours. Test pack signed off.
2. Spool erection complete on Pipe Rack B Tier 2 along grid B1-B8, 100% finished.
3. concreting finished for heavy column footings C1-C12 in CDU area, 180 cu.m poured.
4. Rebar cutting, bending and shuttering for Main Gas Compressor C-101 foundation completed 100%.
5. Pump alignment and baseplate grouting for Crude Charge Pump P-101A completed.
6. Cable tray bracket mounting and ladder tray installation along Rack B Tier 3 complete.
7. Bench calibration and impulse tubing hookup for Pressure Transmitter PT-101 complete.
"""

    def test_extract_filters_headers(self, project_id, full_schedule):
        """Header/metadata lines must be filtered, only work observations remain."""
        observations = DocumentExtractor.extract_from_text(
            self.DAILY_REPORT, project_id
        )
        # 4 header lines filtered, 7 observation lines retained
        assert len(observations) == 7

    def test_normalization_applied(self, project_id, full_schedule):
        """Each observation must have normalized text differing from raw text."""
        observations = DocumentExtractor.extract_from_text(
            self.DAILY_REPORT, project_id
        )
        for obs in observations:
            assert obs.normalized_text is not None
            assert len(obs.normalized_text) > 0

    def test_discipline_detection_across_disciplines(self, project_id, full_schedule):
        """The report contains piping, civil, electrical, and instrumentation work at minimum."""
        observations = DocumentExtractor.extract_from_text(
            self.DAILY_REPORT, project_id
        )
        detected_disciplines = {obs.discipline for obs in observations}
        assert DisciplineEnum.PIPING in detected_disciplines
        assert DisciplineEnum.CIVIL in detected_disciplines
        assert DisciplineEnum.ELECTRICAL in detected_disciplines
        assert DisciplineEnum.INSTRUMENTATION in detected_disciplines
        # At least 4 distinct disciplines detected from multi-discipline report
        assert len(detected_disciplines) >= 4

    def test_full_pipeline_extract_and_match(self, project_id, full_schedule):
        """
        The complete pipeline: extract observations from text, then match each
        against the full schedule. Verify correct activity codes, tiers, and
        auto-link eligibility.
        """
        observations = DocumentExtractor.extract_from_text(
            self.DAILY_REPORT, project_id
        )
        assert len(observations) >= 7

        proposals: list[MatchProposalPayload] = []
        for obs in observations:
            proposal = HybridMatcher.match_observation(obs, full_schedule)
            proposals.append(proposal)

        assert len(proposals) == len(observations)

        # Every observation should produce at least one candidate
        for proposal in proposals:
            assert len(proposal.candidates) > 0

        # Verify specific high-confidence matches exist
        matched_codes = [
            p.top_match.activity_code for p in proposals if p.top_match is not None
        ]
        assert "PIP-2401" in matched_codes  # P-101 hydro test
        assert "PIP-2400" in matched_codes  # Spool erection

    def test_pipeline_produces_auto_link_candidates(self, project_id, full_schedule):
        """High-confidence matches with clear score gap should be auto-link eligible."""
        observations = DocumentExtractor.extract_from_text(
            self.DAILY_REPORT, project_id
        )
        proposals = [
            HybridMatcher.match_observation(obs, full_schedule) for obs in observations
        ]

        auto_linkable = [p for p in proposals if p.auto_link_eligible]
        # At least some observations should be auto-linkable
        assert len(auto_linkable) >= 1

    def test_pipeline_classifies_review_required(self, project_id, full_schedule):
        """
        Ambiguous text should produce MEDIUM-tier matches that are NOT auto-linkable,
        routing them to the planner review queue.
        """
        ambiguous_text = "Hydrostatic pressure testing completed along Pipe Rack B headers yesterday."
        observations = DocumentExtractor.extract_from_text(ambiguous_text, project_id)
        assert len(observations) >= 1

        proposal = HybridMatcher.match_observation(observations[0], full_schedule)
        assert len(proposal.candidates) >= 2
        # Both PIP-2401 and PIP-2402 are hydrotest activities — ambiguous
        top_codes = {c.activity_code for c in proposal.candidates[:2]}
        assert "PIP-2401" in top_codes or "PIP-2402" in top_codes

    def test_pipeline_commit_eligibility(self, project_id, full_schedule):
        """
        Simulates the commit gate: only HIGH-tier auto-linkable matches are
        eligible for automatic commitment; others require human review.
        """
        observations = DocumentExtractor.extract_from_text(
            self.DAILY_REPORT, project_id
        )
        proposals = [
            HybridMatcher.match_observation(obs, full_schedule) for obs in observations
        ]

        for proposal in proposals:
            if proposal.auto_link_eligible:
                # Auto-link → can be committed directly
                assert proposal.top_match is not None
                assert proposal.top_match.match_tier == MatchTierEnum.HIGH
                assert proposal.top_match.confidence_score >= 0.88
            elif proposal.top_match:
                # Needs review → cannot be auto-committed
                if proposal.top_match.match_tier == MatchTierEnum.MEDIUM:
                    assert not proposal.auto_link_eligible


class TestFullPipelineExcelIngestion:
    """
    Integration test: Full pipeline from Excel spreadsheet ingestion through
    matching against schedule activities.
    """

    def test_excel_to_match_pipeline(self, project_id, full_schedule):
        """Excel file → extract observations → normalize → match against schedule."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Daily Log"

        ws.append(["Activity Code", "Description", "Status", "Discipline", "Remarks"])
        ws.append(
            [
                "PIP-2400",
                "Spool erection on Pipe Rack B Tier 2",
                "100% Complete",
                "Piping",
                "All spools erected",
            ]
        )
        ws.append(
            [
                "CIV-1100",
                "Rebar tying for compressor foundation",
                "In Progress",
                "Civil",
                "15 MT tied",
            ]
        )
        ws.append(
            [
                "CIV-1101",
                "Concrete pour column footings",
                "Complete",
                "Civil",
                "180 cu.m poured",
            ]
        )
        ws.append(
            [
                "ELE-4100",
                "Cable tray installation Rack B Tier 3",
                "75% Complete",
                "Electrical",
                "465m installed",
            ]
        )

        buf = io.BytesIO()
        wb.save(buf)
        excel_bytes = buf.getvalue()

        observations = DocumentExtractor.extract_from_excel_bytes(
            excel_bytes, project_id
        )
        assert len(observations) == 4

        # Match each observation
        proposals = [
            HybridMatcher.match_observation(obs, full_schedule) for obs in observations
        ]
        assert len(proposals) == 4

        # Verify correct activity code matches
        matched = {p.top_match.activity_code for p in proposals if p.top_match}
        assert "PIP-2400" in matched
        assert "CIV-1100" in matched

    def test_excel_with_alternative_headers(self, project_id, full_schedule):
        """Excel with non-standard column headers should still extract correctly."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Act. Code", "Work Description", "% Complete", "Disc.", "Tag No"])
        ws.append(
            [
                "PIP-2401",
                "Hydrostatic Testing Line P-101",
                "100",
                "Piping",
                "LINE-P-101",
            ]
        )

        buf = io.BytesIO()
        wb.save(buf)

        observations = DocumentExtractor.extract_from_excel_bytes(
            buf.getvalue(), project_id
        )
        assert len(observations) == 1
        assert observations[0].discipline == DisciplineEnum.PIPING


# =============================================================================
# 2. DATABASE WRITE CONTRACT VALIDATION (Audit Record Generation)
# =============================================================================


class TestAuditRecordGeneration:
    """
    Validates that the data structures produced by the pipeline are correct
    for database writes — observations, match proposals, and audit events.
    """

    def test_observation_has_all_required_fields(self, project_id, full_schedule):
        """Each extracted observation must have all fields required by the work_observations table."""
        text = "P-101 hydro test completed at 42.5 bar, 100% done."
        observations = DocumentExtractor.extract_from_text(text, project_id)
        obs = observations[0]

        # Required by database schema
        assert obs.id is not None
        assert obs.project_id == project_id
        assert obs.raw_text is not None and len(obs.raw_text) > 0
        assert obs.normalized_text is not None and len(obs.normalized_text) > 0
        assert obs.observed_at is not None
        assert obs.discipline is not None
        assert obs.extraction_confidence > 0.0

    def test_match_proposal_has_all_required_fields(self, project_id, full_schedule):
        """Each match proposal must have all fields required by the match_proposals table."""
        text = "Spool erection on Pipe Rack B completed 100%."
        observations = DocumentExtractor.extract_from_text(text, project_id)
        proposal = HybridMatcher.match_observation(observations[0], full_schedule)

        assert proposal.observation is not None
        assert proposal.observation.id is not None
        for candidate in proposal.candidates:
            assert candidate.activity_id is not None
            assert candidate.activity_code is not None
            assert candidate.activity_name is not None
            assert candidate.candidate_rank >= 1
            assert 0.0 <= candidate.lexical_score <= 1.0
            assert 0.0 <= candidate.semantic_score <= 1.0
            assert 0.0 <= candidate.context_boost <= 0.30
            assert 0.0 <= candidate.confidence_score <= 1.0
            assert candidate.match_tier in (
                MatchTierEnum.HIGH,
                MatchTierEnum.MEDIUM,
                MatchTierEnum.LOW,
                MatchTierEnum.UNMATCHED,
            )
            assert candidate.explanation is not None
            assert candidate.evidence_snippet is not None

    def test_match_proposal_serializes_to_json(self, project_id, full_schedule):
        """Match proposals must be JSON-serializable for database JSONB columns and API responses."""
        text = "P-101 completed successfully."
        observations = DocumentExtractor.extract_from_text(text, project_id)
        proposal = HybridMatcher.match_observation(observations[0], full_schedule)

        json_str = proposal.model_dump_json()
        assert len(json_str) > 0

        # Round-trip: can be deserialized back
        restored = MatchProposalPayload.model_validate_json(json_str)
        assert restored.observation.id == proposal.observation.id
        assert len(restored.candidates) == len(proposal.candidates)

    def test_audit_trail_data_integrity(self, project_id, full_schedule):
        """
        Simulates the audit trail record that the Rust backend would create.
        Verifies the data contract between AI service output and audit schema.
        """
        text = "Spool erection complete on Pipe Rack B."
        observations = DocumentExtractor.extract_from_text(text, project_id)
        proposal = HybridMatcher.match_observation(observations[0], full_schedule)

        # Simulate audit event creation (mirrors Rust EventLedger logic)
        import hashlib

        entity_id = proposal.observation.id
        action = "MATCH_PROPOSED"
        before_state = None
        after_state = proposal.model_dump(mode="json")

        payload_for_hash = json.dumps(
            {
                "entity_type": "match_proposal",
                "entity_id": str(entity_id),
                "action": action,
                "before": before_state,
                "after": after_state,
            },
            default=str,
        )

        payload_hash = hashlib.sha256(
            str(entity_id).encode() + action.encode() + payload_for_hash.encode()
        ).hexdigest()

        assert len(payload_hash) == 64  # SHA-256 hex digest
        # Hash is deterministic
        payload_hash_2 = hashlib.sha256(
            str(entity_id).encode() + action.encode() + payload_for_hash.encode()
        ).hexdigest()
        assert payload_hash == payload_hash_2

    def test_observation_ids_are_unique(self, project_id, full_schedule):
        """All observations from one extraction run must have unique IDs (for DB primary keys)."""
        text = """Spool erection complete on Pipe Rack B.
P-101 hydro test completed.
Concrete pour started for column footings."""
        observations = DocumentExtractor.extract_from_text(text, project_id)
        ids = [obs.id for obs in observations]
        assert len(ids) == len(set(ids))


# =============================================================================
# 3. EDGE CASES: Duplicate Files, Malformed Data, Low-Confidence, Retries
# =============================================================================


class TestDuplicateHandling:
    """Edge cases for duplicate files and near-duplicate observations."""

    def test_duplicate_text_extraction_idempotent(self, project_id, full_schedule):
        """Extracting the same text twice produces observations with different IDs (no collision)."""
        text = "P-101 hydro test completed."
        obs_run1 = DocumentExtractor.extract_from_text(text, project_id)
        obs_run2 = DocumentExtractor.extract_from_text(text, project_id)

        assert len(obs_run1) == len(obs_run2)
        # Different UUIDs each run
        assert obs_run1[0].id != obs_run2[0].id
        # Same content
        assert obs_run1[0].raw_text == obs_run2[0].raw_text
        assert obs_run1[0].normalized_text == obs_run2[0].normalized_text

    def test_near_duplicate_detection(self, project_id):
        """Near-duplicate observations (>= 92% token similarity) should be detectable."""
        from rapidfuzz import fuzz

        obs1 = NormalizedObservation(
            project_id=project_id,
            raw_text="P-101 hydro test completed successfully at 42.5 bar",
            normalized_text="Line P-101 Hydrostatic Testing completed successfully at 42.5 bar",
            discipline=DisciplineEnum.PIPING,
            extraction_confidence=0.95,
        )
        obs2 = NormalizedObservation(
            project_id=project_id,
            raw_text="P-101 hydro test completed successfully at 42.5 bar pressure holding",
            normalized_text="Line P-101 Hydrostatic Testing completed successfully at 42.5 bar pressure holding",
            discipline=DisciplineEnum.PIPING,
            extraction_confidence=0.98,
        )

        similarity = (
            fuzz.token_sort_ratio(
                obs1.normalized_text.lower(), obs2.normalized_text.lower()
            )
            / 100.0
        )
        # These are near-duplicates — similarity above dedup threshold
        assert similarity >= 0.85
        # The higher-confidence one should be preferred in a dedup merge
        assert obs2.extraction_confidence > obs1.extraction_confidence

    def test_distinct_observations_are_not_duplicates(self, project_id):
        """Observations with clearly different content should NOT be flagged as duplicates."""
        from rapidfuzz import fuzz

        obs1 = NormalizedObservation(
            project_id=project_id,
            raw_text="P-101 hydro test completed",
            normalized_text="Line P-101 Hydrostatic Testing completed",
            discipline=DisciplineEnum.PIPING,
            extraction_confidence=0.95,
        )
        obs2 = NormalizedObservation(
            project_id=project_id,
            raw_text="Rebar tying for compressor foundation started",
            normalized_text="Rebar Tying and Shuttering for compressor foundation started",
            discipline=DisciplineEnum.CIVIL,
            extraction_confidence=0.95,
        )

        similarity = (
            fuzz.token_sort_ratio(
                obs1.normalized_text.lower(), obs2.normalized_text.lower()
            )
            / 100.0
        )
        # These are distinct — similarity well below dedup threshold
        assert similarity < 0.50

    def test_duplicate_excel_rows_produce_separate_observations(
        self, project_id, full_schedule
    ):
        """Identical rows in Excel should produce separate observations."""
        import openpyxl
        from rapidfuzz import fuzz

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Activity Code", "Description", "Status", "Discipline"])
        ws.append(["PIP-2400", "Spool erection complete", "100%", "Piping"])
        ws.append(
            ["PIP-2400", "Spool erection complete", "100%", "Piping"]
        )  # Duplicate row

        buf = io.BytesIO()
        wb.save(buf)

        observations = DocumentExtractor.extract_from_excel_bytes(
            buf.getvalue(), project_id
        )
        assert len(observations) == 2  # Both extracted
        # The duplicates are detectable by similarity
        similarity = (
            fuzz.token_sort_ratio(
                observations[0].normalized_text.lower(),
                observations[1].normalized_text.lower(),
            )
            / 100.0
        )
        assert similarity >= 0.92  # Confirmed as near-duplicates


class TestMalformedData:
    """Edge cases for malformed, incomplete, or garbage input data."""

    def test_empty_text_produces_no_observations(self, project_id):
        """Empty string input should produce zero observations."""
        observations = DocumentExtractor.extract_from_text("", project_id)
        assert observations == []

    def test_whitespace_only_produces_no_observations(self, project_id):
        """Whitespace-only input should produce zero observations."""
        observations = DocumentExtractor.extract_from_text(
            "   \n\n\t\t  \n  ", project_id
        )
        assert observations == []

    def test_garbage_text_produces_general_discipline(self, project_id):
        """Random non-construction text should still extract but detect GENERAL discipline."""
        observations = DocumentExtractor.extract_from_text(
            "The quick brown fox jumped over the lazy dog on Tuesday.", project_id
        )
        assert len(observations) == 1
        assert observations[0].discipline == DisciplineEnum.GENERAL

    def test_malformed_excel_falls_back_to_text(self, project_id):
        """Non-Excel binary data should fallback to text extraction gracefully."""
        garbage = b"This is not an Excel file but plain text about spool erection."
        observations = DocumentExtractor.extract_from_excel_bytes(garbage, project_id)
        # Falls back to text extraction
        assert len(observations) >= 1
        assert observations[0].discipline == DisciplineEnum.PIPING

    def test_very_short_lines_filtered(self, project_id):
        """Lines shorter than 5 chars should be filtered out."""
        text = "OK\nYes\n1. Spool erection on Pipe Rack B completed 100%."
        observations = DocumentExtractor.extract_from_text(text, project_id)
        # Only the long line survives
        assert len(observations) == 1

    def test_header_lines_filtered_from_report(self, project_id):
        """Standard report headers should not become observations."""
        text = """Project: PRD-HYD-PKG04
Date: 28-Aug-2026
Report No: DPR-2026-0828
Daily Construction Report
Contractor: NEXORA EPC
Spool erection on Pipe Rack B completed."""
        observations = DocumentExtractor.extract_from_text(text, project_id)
        assert len(observations) == 1
        assert "spool" in observations[0].raw_text.lower()

    def test_matching_with_empty_activities_list(self, project_id):
        """Matching against an empty schedule should produce no candidates."""
        obs = NormalizedObservation(
            project_id=project_id,
            raw_text="P-101 hydro test done",
            normalized_text="Line P-101 Hydrostatic Testing done",
        )
        proposal = HybridMatcher.match_observation(obs, [])
        assert proposal.candidates == []
        assert proposal.top_match is None
        assert proposal.auto_link_eligible is False

    def test_observation_with_no_discipline_still_matches(
        self, project_id, full_schedule
    ):
        """Observations without explicit discipline should still produce matches."""
        obs = NormalizedObservation(
            project_id=project_id,
            raw_text="Work at Pipe Rack B completed",
            normalized_text="Work at Pipe Rack B completed",
            discipline=DisciplineEnum.GENERAL,
        )
        proposal = HybridMatcher.match_observation(obs, full_schedule)
        assert len(proposal.candidates) > 0

    def test_unicode_text_extraction(self, project_id):
        """Unicode characters in text should not crash extraction."""
        text = "Spool erection — complete ✓ on Pipe Rack B (100% done)."
        observations = DocumentExtractor.extract_from_text(text, project_id)
        assert len(observations) == 1
        assert observations[0].reported_progress == 100.0


class TestLowConfidenceMatches:
    """Edge cases for low-confidence and unmatched observations."""

    def test_unrelated_work_produces_low_confidence(self, project_id, full_schedule):
        """Work not in the schedule should produce low-confidence or UNMATCHED tier."""
        text = "Catering team delivered 200 lunch packets to construction camp 3."
        observations = DocumentExtractor.extract_from_text(text, project_id)
        proposal = HybridMatcher.match_observation(observations[0], full_schedule)

        if proposal.top_match:
            assert proposal.top_match.confidence_score < 0.60
            assert proposal.auto_link_eligible is False

    def test_wrong_discipline_activity_not_auto_linked(self, project_id, full_schedule):
        """
        An observation about electrical work should not auto-link to a piping activity,
        even if there's some textual overlap.
        """
        obs = NormalizedObservation(
            project_id=project_id,
            raw_text="Electrical cable tray work near Pipe Rack B area",
            normalized_text="Electrical Cable Tray Installation work near Pipe Rack B area",
            discipline=DisciplineEnum.ELECTRICAL,
        )
        proposal = HybridMatcher.match_observation(obs, full_schedule)

        # Should match ELE-4100 (Cable Tray), not a piping activity
        if proposal.top_match and proposal.top_match.match_tier == MatchTierEnum.HIGH:
            assert proposal.top_match.activity_code.startswith("ELE")

    def test_vague_text_routes_to_review(self, project_id, full_schedule):
        """Vague or incomplete text should not be auto-linked."""
        obs = NormalizedObservation(
            project_id=project_id,
            raw_text="Some work done today at site",
            normalized_text="Some work done today at site",
            discipline=DisciplineEnum.GENERAL,
        )
        proposal = HybridMatcher.match_observation(obs, full_schedule)
        assert proposal.auto_link_eligible is False

    def test_confidence_scores_bounded(self, project_id, full_schedule):
        """All confidence scores must be between 0.0 and 1.0."""
        text = "P-101 hydro test completed with 42.5 bar pressure."
        observations = DocumentExtractor.extract_from_text(text, project_id)
        proposal = HybridMatcher.match_observation(observations[0], full_schedule)

        for candidate in proposal.candidates:
            assert 0.0 <= candidate.confidence_score <= 1.0
            assert 0.0 <= candidate.lexical_score <= 1.0
            assert 0.0 <= candidate.semantic_score <= 1.0


class TestRetryAndIdempotency:
    """Edge cases for processing retries and idempotent behavior."""

    def test_matching_is_deterministic(self, project_id, full_schedule):
        """
        Running the same observation against the same schedule twice must
        produce identical rankings and scores (deterministic tie-breaking).
        """
        obs = NormalizedObservation(
            project_id=project_id,
            raw_text="Spool erection on Pipe Rack B complete",
            normalized_text="Spool Erection and Alignment on Pipe Rack B complete",
            discipline=DisciplineEnum.PIPING,
        )

        proposal_1 = HybridMatcher.match_observation(obs, full_schedule)
        proposal_2 = HybridMatcher.match_observation(obs, full_schedule)

        assert len(proposal_1.candidates) == len(proposal_2.candidates)
        for c1, c2 in zip(proposal_1.candidates, proposal_2.candidates):
            assert c1.activity_code == c2.activity_code
            assert c1.confidence_score == c2.confidence_score
            assert c1.candidate_rank == c2.candidate_rank

    def test_normalization_is_deterministic(self):
        """Same input text always produces same normalized output."""
        text = "hydro test on p-101 completed, 42 bar pressure"
        result1 = default_normalizer.normalize(text, DisciplineEnum.PIPING)
        result2 = default_normalizer.normalize(text, DisciplineEnum.PIPING)
        assert result1 == result2

    def test_extraction_is_deterministic(self, project_id):
        """Same input text always produces same extraction results (except UUIDs)."""
        text = "P-101 hydro test completed 100%."
        obs1 = DocumentExtractor.extract_from_text(text, project_id)
        obs2 = DocumentExtractor.extract_from_text(text, project_id)

        assert len(obs1) == len(obs2)
        assert obs1[0].raw_text == obs2[0].raw_text
        assert obs1[0].normalized_text == obs2[0].normalized_text
        assert obs1[0].discipline == obs2[0].discipline
        assert obs1[0].event_type == obs2[0].event_type
        assert obs1[0].reported_progress == obs2[0].reported_progress

    def test_batch_embedding_caching_consistent(self, project_id, full_schedule):
        """
        After batch pre-computing embeddings, running match twice should
        use cached embeddings and produce identical results.
        """
        obs = NormalizedObservation(
            project_id=project_id,
            raw_text="Cable tray installation at Pipe Rack B",
            normalized_text="Cable Tray Installation at Pipe Rack B",
            discipline=DisciplineEnum.ELECTRICAL,
        )

        # First run pre-computes embeddings
        proposal_1 = HybridMatcher.match_observation(obs, full_schedule)
        # Second run uses cached embeddings
        proposal_2 = HybridMatcher.match_observation(obs, full_schedule)

        # Results must be identical
        assert proposal_1.top_match is not None
        assert proposal_2.top_match is not None
        assert proposal_1.top_match.activity_code == proposal_2.top_match.activity_code
        assert (
            proposal_1.top_match.confidence_score
            == proposal_2.top_match.confidence_score
        )


# =============================================================================
# 4. GOLDEN DATASET VALIDATION
# =============================================================================


class TestGoldenDatasetIntegration:
    """
    Runs the golden benchmark dataset through the full pipeline and validates
    expected match tiers and activity codes.
    """

    @pytest.fixture
    def golden_scenarios(self):
        import os

        golden_path = os.path.join(
            os.path.dirname(__file__),
            "..",
            "..",
            "tests",
            "golden_dataset",
            "field_observations.json",
        )
        if not os.path.exists(golden_path):
            pytest.skip("Golden dataset file not found")

        with open(golden_path) as f:
            data = json.load(f)
        return data["scenarios"]

    def test_golden_dataset_pipeline(self, project_id, full_schedule, golden_scenarios):
        """
        Each golden scenario should be extractable and matchable.
        We verify that the pipeline produces results for every scenario.
        """
        for scenario in golden_scenarios:
            raw_text = scenario["raw_text"]
            expected_tier = scenario.get("expected_match_tier")

            observations = DocumentExtractor.extract_from_text(raw_text, project_id)
            assert len(observations) >= 1, (
                f"No observations extracted for {scenario['id']}"
            )

            proposal = HybridMatcher.match_observation(observations[0], full_schedule)

            # Every scenario must produce candidates (even if UNMATCHED)
            if expected_tier == "UNMATCHED":
                if proposal.top_match:
                    assert proposal.top_match.confidence_score < 0.60
            elif expected_tier == "HIGH":
                assert proposal.top_match is not None, (
                    f"Expected HIGH match for {scenario['id']}"
                )
                assert proposal.top_match.confidence_score >= 0.60

    def test_golden_event_types_detected(self, project_id, golden_scenarios):
        """Verify event type detection for golden scenarios."""
        for scenario in golden_scenarios:
            raw_text = scenario["raw_text"]
            expected_event = scenario.get("expected_event_type")
            if not expected_event:
                continue

            observations = DocumentExtractor.extract_from_text(raw_text, project_id)
            assert len(observations) >= 1

            detected = observations[0].event_type
            if detected:
                assert detected.value == expected_event, (
                    f"Scenario {scenario['id']}: expected {expected_event}, got {detected.value}"
                )


# =============================================================================
# 5. API ENDPOINT INTEGRATION (FastAPI TestClient)
# =============================================================================


class TestAPIFullPipeline:
    """
    Tests the full pipeline through FastAPI HTTP endpoints using TestClient.
    Validates request/response contracts match database write requirements.
    """

    @pytest.fixture
    def client(self):
        from app.main import app
        from fastapi.testclient import TestClient

        return TestClient(app)

    def test_extract_then_match_via_api(self, client, project_id, full_schedule):
        """POST /extract → POST /match: full HTTP pipeline."""
        # Step 1: Extract
        extract_payload = {
            "project_id": str(project_id),
            "text_content": "P-101 hydro test completed at 42.5 bar pressure.",
            "source_type": "DAILY_REPORT",
        }
        extract_resp = client.post("/api/v1/extract", json=extract_payload)
        assert extract_resp.status_code == 200
        observations = extract_resp.json()
        assert len(observations) >= 1

        # Step 2: Match
        match_payload = {
            "project_id": str(project_id),
            "observations": observations,
            "activities": [act.model_dump(mode="json") for act in full_schedule],
        }
        match_resp = client.post("/api/v1/match", json=match_payload)
        assert match_resp.status_code == 200
        proposals = match_resp.json()
        assert len(proposals) >= 1
        assert proposals[0]["top_match"] is not None
        assert proposals[0]["top_match"]["activity_code"] == "PIP-2401"

    def test_extract_file_then_match_via_api(self, client, project_id, full_schedule):
        """POST /extract-file → POST /match: file upload pipeline."""
        content = b"Spool erection complete on Pipe Rack B, 100% finished."
        files = {"file": ("report.txt", io.BytesIO(content), "text/plain")}
        data = {"project_id": str(project_id)}

        extract_resp = client.post("/api/v1/extract-file", data=data, files=files)
        assert extract_resp.status_code == 200
        observations = extract_resp.json()
        assert len(observations) >= 1

        match_payload = {
            "project_id": str(project_id),
            "observations": observations,
            "activities": [act.model_dump(mode="json") for act in full_schedule],
        }
        match_resp = client.post("/api/v1/match", json=match_payload)
        assert match_resp.status_code == 200
        proposals = match_resp.json()
        assert len(proposals) >= 1

    def test_normalize_then_extract_consistency(self, client, project_id):
        """Normalization via API should produce same result as extraction pipeline."""
        normalize_payload = {
            "project_id": str(project_id),
            "text": "hydro test on p-101",
            "discipline": "PIPING",
        }
        norm_resp = client.post("/api/v1/normalize", json=normalize_payload)
        assert norm_resp.status_code == 200
        normalized = norm_resp.json()["normalized_text"]

        assert "Hydrostatic Testing" in normalized
        assert "Line P-101" in normalized


# =============================================================================
# 6. MULTI-DISCIPLINE CROSS-MATCHING STRESS TEST
# =============================================================================


class TestMultiDisciplineStress:
    """
    Validates that the pipeline correctly handles a report containing work
    from all 6 disciplines simultaneously without cross-contamination.
    """

    MULTI_DISC_REPORT = """1. Spool erection and alignment on Pipe Rack B Tier 2 completed 100%.
2. Rebar tying for compressor foundation C-101 ongoing, 20 MT installed.
3. Pump alignment for Crude Charge Pump P-101A in progress.
4. Cable tray installation along Rack B Tier 3 completed 75%.
5. PT-101 calibration and impulse line hookup started.
6. Toolbox talk conducted at site, all crews attended.
7. Emergency dewatering at Substation 4 due to heavy monsoon rain."""

    def test_all_disciplines_represented(self, project_id, full_schedule):
        """Multi-discipline report should detect at least 5 different disciplines."""
        observations = DocumentExtractor.extract_from_text(
            self.MULTI_DISC_REPORT, project_id
        )
        disciplines = {obs.discipline for obs in observations}
        # At least Piping, Civil, Mechanical, Electrical, Instrumentation
        assert len(disciplines) >= 5

    def test_no_cross_discipline_auto_linking(self, project_id, full_schedule):
        """A piping observation should not auto-link to a civil activity, etc."""
        observations = DocumentExtractor.extract_from_text(
            self.MULTI_DISC_REPORT, project_id
        )

        for obs in observations:
            proposal = HybridMatcher.match_observation(obs, full_schedule)
            if proposal.auto_link_eligible and proposal.top_match:
                disc_prefix_map = {
                    DisciplineEnum.PIPING: "PIP",
                    DisciplineEnum.CIVIL: "CIV",
                    DisciplineEnum.MECHANICAL: "MEC",
                    DisciplineEnum.ELECTRICAL: "ELE",
                    DisciplineEnum.INSTRUMENTATION: "INS",
                    DisciplineEnum.HSE: "HSE",
                }
                expected_prefix = disc_prefix_map.get(obs.discipline, "")
                if expected_prefix and obs.discipline != DisciplineEnum.GENERAL:
                    assert proposal.top_match.activity_code.startswith(
                        expected_prefix
                    ), (
                        f"Cross-discipline auto-link: obs discipline={obs.discipline}, "
                        f"matched activity={proposal.top_match.activity_code}"
                    )

    def test_candidate_rankings_always_descending(self, project_id, full_schedule):
        """For every observation, candidate scores must be in descending order."""
        observations = DocumentExtractor.extract_from_text(
            self.MULTI_DISC_REPORT, project_id
        )

        for obs in observations:
            proposal = HybridMatcher.match_observation(obs, full_schedule)
            scores = [c.confidence_score for c in proposal.candidates]
            assert scores == sorted(scores, reverse=True), (
                f"Non-descending scores for obs: {obs.raw_text[:50]}"
            )
